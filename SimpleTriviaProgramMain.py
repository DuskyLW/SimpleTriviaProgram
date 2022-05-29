# from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import random

wb = load_workbook(filename="questions.xlsx")
ws = wb.active

# Call this to have the question you want printed on screen


def questionprinter(question):
    for key, value in question.items():
        if "question" in key:
            print(value)
        elif "answers" in key:
            for option in value:
                print(option)
        else:
            break

# Call this to get input from the user for the question


def getuseranswer():
    answers = ["A", "B", "C", "D"]
    while True:
        useranswer = str(input("Your answer: ")).upper()
        if useranswer not in answers:
            print("Please input a valid answer of A, B, C or D.\n")
            continue
        break
    return answers.index(useranswer)

# Call this to take user's answer and check it against the question


def answerchecker(answer, question):
    usercorrect = False
    correctanswer = int((question["correctanswer"] - 1))
    questionanswers = question["answers"]
    if answer == correctanswer:
        print("\nCorrect!")
        usercorrect = True
    else:
        print("\nWrong!")
        print("The correct answer was:", str(questionanswers[correctanswer]))
    return usercorrect
# Call this to take user's answer and check it against the question


def databasequestiongrabber():
    qlist = []
    qdictionary = {}
    for row in [random.randint(2, ws.max_row)]:
        for col in range(1, 7):
            char = get_column_letter(col)
            qlist.append(ws[char + str(row)].value)
    qdictionary.update({"question": str("\n" + qlist[0])})
    qdictionary.update({"answers": qlist[1:5]})
    qdictionary.update({"correctanswer": int(qlist[5])})
    return qdictionary

# Main function, do everything here plz


def main():
    print("Welcome to my Trivia program! Let's get started.")
    amountwrong = 0
    amountcorrect = 0
    while True:
        question = databasequestiongrabber()
        questionprinter(question)
        correctorwrong = answerchecker(getuseranswer(), question)
        if not correctorwrong:
            amountwrong += 1
        else:
            amountcorrect += 1
        print("\nAnswered Correct:", str(amountcorrect),
              "Answered Wrong:", str(amountwrong))


if __name__ == "__main__":
    main()
