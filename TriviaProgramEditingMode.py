from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import random

wb = load_workbook(filename="questions.xlsx")
ws = wb.active
db_row_ind = 1
db_column_ind = 1

for row in [2, ws.max_row]:
    print("\nDatabase Row #" + str(db_row_ind))
    for col in range(1, 7):
        sslist = []
        char = get_column_letter(col)
        sslist.append(ws[char + str(row)].value)
        print("Current Row Column #" + str(db_column_ind), str(sslist))
        db_column_ind += 1
    db_row_ind += 1
    db_column_ind = 1

db_row_ind = 1
print(db_row_ind)
print(db_column_ind)

print("Hello!")
